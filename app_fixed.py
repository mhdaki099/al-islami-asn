import streamlit as st
import pandas as pd
import openai
import os
from dotenv import load_dotenv
import PyPDF2
import pdfplumber
from PIL import Image
import fitz  # PyMuPDF
import io
import re
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import traceback

# Load environment variables
load_dotenv()

# Configure page
st.set_page_config(
    page_title="Invoice Data Extractor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize OpenAI
def initialize_openai():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or api_key == "your_openai_api_key_here":
        st.error("Please set your OPENAI_API_KEY in the .env file")
        st.stop()
    openai.api_key = api_key
    return openai

# Enhanced text extraction with better file handling
def extract_text_from_pdf(pdf_file):
    text = ""
    methods_used = []
    
    try:
        # Get file content as bytes
        pdf_file.seek(0)
        pdf_bytes = pdf_file.read()
        
        if not pdf_bytes:
            st.error("PDF file is empty or corrupted")
            return ""
        
        st.info(f"PDF file size: {len(pdf_bytes)} bytes")
        
        # Method 1: Try pdfplumber (best for searchable PDFs)
        try:
            pdf_file.seek(0)
            with pdfplumber.open(pdf_file) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text += page_text + "\n"
                if text.strip():
                    methods_used.append("pdfplumber")
        except Exception as e:
            st.warning(f"pdfplumber failed: {str(e)}")
        
        # Method 2: Try PyPDF2 as fallback
        if not text.strip():
            try:
                pdf_file.seek(0)
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text += page_text + "\n"
                if text.strip():
                    methods_used.append("PyPDF2")
            except Exception as e:
                st.warning(f"PyPDF2 failed: {str(e)}")
        
        # Method 3: OCR with PyMuPDF for scanned PDFs (only if tesseract is available)
        if not text.strip():
            try:
                import pytesseract
                # Use bytes directly instead of file stream
                pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
                for page_num in range(pdf_document.page_count):
                    page = pdf_document[page_num]
                    # Convert page to image with higher resolution
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better OCR
                    pix = page.get_pixmap(matrix=mat)
                    img_data = pix.tobytes("png")
                    image = Image.open(io.BytesIO(img_data))
                    
                    # Use OCR with different PSM modes
                    for psm in [6, 3, 4]:  # Try different page segmentation modes
                        try:
                            page_text = pytesseract.image_to_string(image, config=f'--psm {psm}')
                            if page_text and page_text.strip():
                                text += page_text + "\n"
                                break
                        except:
                            continue
                
                pdf_document.close()
                if text.strip():
                    methods_used.append("OCR (PyMuPDF + Tesseract)")
            except ImportError:
                st.info("🔍 OCR not available locally. Scanned PDFs will work on Streamlit Cloud.")
            except Exception as e:
                st.warning(f"OCR failed: {str(e)}")
        
        # Method 4: Try alternative PyMuPDF text extraction
        if not text.strip():
            try:
                pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
                for page_num in range(pdf_document.page_count):
                    page = pdf_document[page_num]
                    page_text = page.get_text()
                    if page_text and page_text.strip():
                        text += page_text + "\n"
                pdf_document.close()
                if text.strip():
                    methods_used.append("PyMuPDF Text Extraction")
            except Exception as e:
                st.warning(f"PyMuPDF text extraction failed: {str(e)}")
        
        if text.strip():
            st.success(f"Text extracted using: {', '.join(methods_used)}")
            st.info(f"Extracted text length: {len(text)} characters")
        else:
            st.error("Could not extract text using any method")
            st.info("💡 This might be a scanned PDF or corrupted file. OCR will work on Streamlit Cloud!")
            
    except Exception as e:
        st.error(f"Error extracting text from PDF: {str(e)}")
        st.error(f"Traceback: {traceback.format_exc()}")
        return ""
    
    return text

# Enhanced data extraction with better prompt engineering
def extract_invoice_data(text, openai_client, file_name=""):
    # Clean and preprocess text
    text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
    text = text.strip()
    
    if len(text) < 50:
        st.warning(f"Text too short for reliable extraction: {len(text)} characters")
        return None
    
    prompt = f"""
    You are an expert at extracting structured data from invoices. Extract the following information from the provided invoice text and return it in JSON format.
    If any information is not found, use "N/A" as the value.
    
    IMPORTANT: Look for these fields with various possible names/abbreviations:
    - PO Number (could be: PO No, Purchase Order, P.O. Number, Order No, etc.)
    - Item Code (could be: Item No, Product Code, SKU, Part Number, Product ID, etc.)
    - Description (could be: Product Description, Item Description, Product Name, etc.)
    - UOM (could be: Unit of Measure, Unit, U/M, Unit Type, etc.)
    - Quantity (could be: Qty, Amount, Qty Ordered, etc.)
    - Lot Number (could be: Lot No, Batch Number, Batch No, Lot ID, etc.)
    - Expiry Date (could be: Exp Date, Expiration Date, Use By Date, etc.)
    - Mfg Date (could be: Manufacturing Date, Mfg Date, Production Date, Made Date, etc.)
    - Invoice No (could be: Invoice Number, Inv No, Invoice ID, etc.)
    - Unit Price (could be: Price per Unit, Unit Cost, Price, Rate, etc.)
    - Total Price (could be: Line Total, Item Total, Amount, etc.)
    - Country (could be: Origin Country, Country of Origin, Made In, etc.)
    - HS Code (could be: HSN Code, Tariff Code, Customs Code, etc.)
    - Date of Invoice (could be: Invoice Date, Date, Issue Date, etc.)
    - Customer No (could be: Customer Number, Customer ID, Client No, Account No, etc.)
    - Payer Name (could be: Payer, Bill To, Billing Name, etc.)
    - Currency (could be: Curr, Currency Code, etc.)
    - Supplier Name (could be: Vendor Name, Supplier, Company Name, Seller, etc.)
    - Total Amount of the Invoice (could be: Grand Total, Total Amount, Invoice Total, Net Total, etc.)
    - Total VAT or Tax (could be: VAT, Tax, Tax Amount, VAT Amount, Tax Total, etc.)
    
    Instructions:
    1. Look carefully through the entire text
    2. Extract numerical values as numbers (not strings) when possible
    3. Extract dates in a consistent format (YYYY-MM-DD if possible)
    4. Be flexible with field names and variations
    5. If multiple items are present, extract the first/main item or aggregate data
    
    Invoice text:
    {text}
    
    Return only valid JSON format with the above fields as keys. Use the exact field names provided above.
    """
    
    try:
        response = openai_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are an expert at extracting structured data from invoices. Always return valid JSON format with the exact field names provided. Be thorough and accurate."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2000
        )
        
        return response.choices[0].message.content
    except Exception as e:
        st.error(f"Error calling OpenAI API: {str(e)}")
        return None

# Enhanced Excel file creation with better formatting
def create_excel_file(data_list, filename="extracted_invoice_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    
    # Define headers
    headers = [
        "PO Number", "Item Code", "Description", "UOM", "Quantity", "Lot Number",
        "Expiry Date", "Mfg Date", "Invoice No", "Unit Price", "Total Price",
        "Country", "HS Code", "Date of Invoice", "Customer No", "Payer Name",
        "Currency", "Supplier Name", "Total Amount of the Invoice", "Total VAT or Tax"
    ]
    
    # Add headers with enhanced formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data with alternating row colors
    data_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row, data in enumerate(data_list, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=data.get(header, "N/A"))
            if row % 2 == 0:
                cell.fill = data_fill_even
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Save file
    wb.save(filename)
    return filename

# Main Streamlit app
def main():
    st.title("📄 Invoice Data Extractor")
    st.markdown("Upload one or multiple PDF files to extract invoice data using AI")
    
    # Sidebar for configuration
    with st.sidebar:
        st.header("Configuration")
        st.markdown("### OpenAI Settings")
        api_key = st.text_input("OpenAI API Key", type="password", value=os.getenv("OPENAI_API_KEY", ""))
        if api_key:
            os.environ["OPENAI_API_KEY"] = api_key
        
        st.markdown("### Processing Options")
        max_file_size = st.slider("Max File Size (MB)", 1, 50, 10)
        
        # Show OCR status
        st.markdown("### OCR Status")
        try:
            import pytesseract
            st.success("✅ OCR Available")
            st.info("Can process scanned PDFs")
        except ImportError:
            st.warning("⚠️ OCR Not Available Locally")
            st.info("🔍 OCR will work on Streamlit Cloud")
            st.info("📄 Searchable PDFs work locally")
        
        st.markdown("### Tips")
        st.info("💡 **Searchable PDFs**: Work locally\n🔍 **Scanned PDFs**: Work on Streamlit Cloud")
    
    # Initialize OpenAI
    if not api_key:
        st.error("Please enter your OpenAI API key in the sidebar")
        st.stop()
    
    openai_client = initialize_openai()
    
    # File upload
    uploaded_files = st.file_uploader(
        "Choose PDF files",
        type=['pdf'],
        accept_multiple_files=True,
        help="Upload one or multiple PDF files (searchable or scanned)"
    )
    
    if uploaded_files:
        # Check file sizes
        large_files = [f for f in uploaded_files if f.size > max_file_size * 1024 * 1024]
        if large_files:
            st.warning(f"Some files are larger than {max_file_size}MB: {[f.name for f in large_files]}")
        
        st.success(f"Uploaded {len(uploaded_files)} file(s)")
        
        # Process files
        if st.button("Extract Data", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            all_extracted_data = []
            processing_log = []
            
            for i, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"Processing {uploaded_file.name}...")
                
                try:
                    # Extract text from PDF
                    text = extract_text_from_pdf(uploaded_file)
                    
                    if text.strip():
                        # Extract data using OpenAI
                        extracted_data = extract_invoice_data(text, openai_client, uploaded_file.name)
                        
                        if extracted_data:
                            try:
                                # Parse JSON response
                                data_dict = json.loads(extracted_data)
                                data_dict['Source File'] = uploaded_file.name
                                data_dict['Processing Time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                all_extracted_data.append(data_dict)
                                processing_log.append(f"✅ {uploaded_file.name}: Success")
                            except json.JSONDecodeError as e:
                                st.warning(f"Could not parse data from {uploaded_file.name}: {str(e)}")
                                processing_log.append(f"❌ {uploaded_file.name}: JSON Parse Error")
                        else:
                            st.warning(f"Could not extract data from {uploaded_file.name}")
                            processing_log.append(f"❌ {uploaded_file.name}: No Data Extracted")
                    else:
                        st.warning(f"No text could be extracted from {uploaded_file.name}")
                        processing_log.append(f"❌ {uploaded_file.name}: No Text Extracted")
                
                except Exception as e:
                    st.error(f"Error processing {uploaded_file.name}: {str(e)}")
                    processing_log.append(f"❌ {uploaded_file.name}: Error - {str(e)}")
                
                progress_bar.progress((i + 1) / len(uploaded_files))
            
            # Display results
            st.markdown("---")
            st.subheader("Processing Results")
            
            # Show processing log
            with st.expander("Processing Log"):
                for log_entry in processing_log:
                    st.text(log_entry)
            
            if all_extracted_data:
                # Create DataFrame
                df = pd.DataFrame(all_extracted_data)
                
                # Display extracted data
                st.success(f"Successfully extracted data from {len(all_extracted_data)} file(s)!")
                st.dataframe(df, use_container_width=True)
                
                # Create and download Excel file
                excel_filename = f"extracted_invoice_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                create_excel_file(all_extracted_data, excel_filename)
                
                with open(excel_filename, "rb") as file:
                    st.download_button(
                        label="📥 Download Excel File",
                        data=file.read(),
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # Clean up temporary file
                try:
                    os.remove(excel_filename)
                except:
                    pass
            else:
                st.error("No data could be extracted from any of the files")
                st.info("💡 **Tips:**\n- Try uploading searchable PDFs for local processing\n- Scanned PDFs will work on Streamlit Cloud\n- Check if the files contain readable text")

if __name__ == "__main__":
    main()
