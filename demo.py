#!/usr/bin/env python3
"""
Demo script to test the invoice data extraction functionality
"""

import os
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

def test_openai_connection():
    """Test OpenAI API connection"""
    try:
        import openai
        openai.api_key = os.getenv("OPENAI_API_KEY")
        
        # Test with a simple prompt
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "user", "content": "Hello, this is a test. Please respond with 'API connection successful'."}
            ],
            max_tokens=50
        )
        
        print("✅ OpenAI API connection successful")
        print(f"Response: {response.choices[0].message.content}")
        return True
    except Exception as e:
        print(f"❌ OpenAI API connection failed: {e}")
        return False

def test_pdf_processing():
    """Test PDF processing capabilities"""
    try:
        import PyPDF2
        import pdfplumber
        import pytesseract
        from PIL import Image
        import fitz
        
        print("✅ PDF processing libraries imported successfully")
        return True
    except Exception as e:
        print(f"❌ PDF processing libraries import failed: {e}")
        return False

def test_excel_creation():
    """Test Excel file creation"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add some test data
        ws['A1'] = "Test Header"
        ws['A2'] = "Test Data"
        
        # Apply formatting
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Save test file
        test_filename = "test_output.xlsx"
        wb.save(test_filename)
        
        # Clean up
        os.remove(test_filename)
        
        print("✅ Excel file creation successful")
        return True
    except Exception as e:
        print(f"❌ Excel file creation failed: {e}")
        return False

def test_data_extraction():
    """Test data extraction with sample text"""
    try:
        import openai
        openai.api_key = os.getenv("OPENAI_API_KEY")
        
        # Sample invoice text
        sample_text = """
        INVOICE
        
        Invoice No: INV-2024-001
        Date: 2024-01-15
        Supplier: ABC Company Ltd
        Customer: XYZ Corporation
        PO Number: PO-2024-001
        
        Item Code: ITM-001
        Description: Sample Product
        Quantity: 10
        Unit Price: $25.00
        Total Price: $250.00
        Currency: USD
        Country: USA
        HS Code: 1234.56.78
        """
        
        prompt = f"""
        Extract the following information from this invoice text and return it in JSON format.
        If any information is not found, use "N/A" as the value.
        
        Required fields:
        - PO Number
        - Item Code
        - Description
        - UOM
        - Quantity
        - Lot Number
        - Expiry Date
        - Mfg Date
        - Invoice No
        - Unit Price
        - Total Price
        - Country
        - HS Code
        - Date of Invoice
        - Customer No
        - Payer Name
        - Currency
        - Supplier Name
        - Total Amount of the Invoice
        - Total VAT or Tax
        
        Invoice text:
        {sample_text}
        
        Return only valid JSON format with the above fields as keys.
        """
        
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are an expert at extracting structured data from invoices. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=1000
        )
        
        result = response.choices[0].message.content
        print("✅ Data extraction successful")
        print("Sample extraction result:")
        print(result)
        
        # Try to parse JSON
        try:
            data = json.loads(result)
            print("✅ JSON parsing successful")
            return True
        except json.JSONDecodeError:
            print("❌ JSON parsing failed")
            return False
            
    except Exception as e:
        print(f"❌ Data extraction failed: {e}")
        return False

def main():
    """Run all tests"""
    print("🔍 Testing Invoice Data Extractor Demo\n")
    
    tests = [
        ("OpenAI API Connection", test_openai_connection),
        ("PDF Processing Libraries", test_pdf_processing),
        ("Excel File Creation", test_excel_creation),
        ("Data Extraction", test_data_extraction)
    ]
    
    results = []
    for test_name, test_func in tests:
        print(f"\n--- {test_name} ---")
        results.append(test_func())
    
    print(f"\n{'='*50}")
    if all(results):
        print("🎉 All tests passed! Your setup is ready.")
        print("\nTo run the app:")
        print("  streamlit run app.py")
        print("  streamlit run streamlit_app.py  # For Streamlit Cloud")
    else:
        print("❌ Some tests failed. Please fix the issues above.")
        print("\nCommon solutions:")
        print("1. Install missing packages: pip install -r requirements.txt")
        print("2. Set your OpenAI API key in .env file")
        print("3. Install Tesseract OCR on your system")

if __name__ == "__main__":
    main()
