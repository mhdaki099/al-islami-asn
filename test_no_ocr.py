#!/usr/bin/env python3
"""
Test script to verify the setup without OCR dependencies
"""

import sys
import importlib

def test_imports():
    """Test if all required packages can be imported"""
    required_packages = [
        'streamlit',
        'pandas',
        'openai',
        'PyPDF2',
        'pdfplumber',
        'PIL',
        'fitz',
        'openpyxl',
        'dotenv'
    ]
    
    print("Testing package imports...")
    failed_imports = []
    
    for package in required_packages:
        try:
            importlib.import_module(package)
            print(f"✅ {package}")
        except ImportError as e:
            print(f"❌ {package}: {e}")
            failed_imports.append(package)
    
    if failed_imports:
        print(f"\n❌ Failed to import: {', '.join(failed_imports)}")
        print("Please install missing packages with: pip install -r requirements_no_ocr.txt")
        return False
    else:
        print("\n✅ All packages imported successfully!")
        return True

def test_openai_key():
    """Test if OpenAI API key is configured"""
    try:
        from dotenv import load_dotenv
        import os
        
        load_dotenv()
        api_key = os.getenv("OPENAI_API_KEY")
        
        if api_key and api_key != "your_openai_api_key_here":
            print("✅ OpenAI API key is configured")
            return True
        else:
            print("❌ OpenAI API key not found or not configured")
            print("Please set your OPENAI_API_KEY in the .env file")
            return False
    except Exception as e:
        print(f"❌ Error checking OpenAI key: {e}")
        return False

def test_pdf_processing():
    """Test PDF processing capabilities without OCR"""
    try:
        import PyPDF2
        import pdfplumber
        from PIL import Image
        import fitz
        
        print("✅ PDF processing libraries imported successfully")
        print("ℹ️  OCR not available locally - will work on Streamlit Cloud")
        return True
    except Exception as e:
        print(f"❌ PDF processing libraries import failed: {e}")
        return False

def test_excel_creation():
    """Test Excel file creation"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add some test data
        ws['A1'] = "Test Header"
        ws['A2'] = "Test Data"
        
        # Apply formatting
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Save test file
        test_filename = "test_output.xlsx"
        wb.save(test_filename)
        
        # Clean up
        os.remove(test_filename)
        
        print("✅ Excel file creation successful")
        return True
    except Exception as e:
        print(f"❌ Excel file creation failed: {e}")
        return False

def test_data_extraction():
    """Test data extraction with sample text"""
    try:
        import openai
        openai.api_key = os.getenv("OPENAI_API_KEY")
        
        # Sample invoice text
        sample_text = """
        INVOICE
        
        Invoice No: INV-2024-001
        Date: 2024-01-15
        Supplier: ABC Company Ltd
        Customer: XYZ Corporation
        PO Number: PO-2024-001
        
        Item Code: ITM-001
        Description: Sample Product
        Quantity: 10
        Unit Price: $25.00
        Total Price: $250.00
        Currency: USD
        Country: USA
        HS Code: 1234.56.78
        """
        
        prompt = f"""
        Extract the following information from this invoice text and return it in JSON format.
        If any information is not found, use "N/A" as the value.
        
        Required fields:
        - PO Number
        - Item Code
        - Description
        - UOM
        - Quantity
        - Lot Number
        - Expiry Date
        - Mfg Date
        - Invoice No
        - Unit Price
        - Total Price
        - Country
        - HS Code
        - Date of Invoice
        - Customer No
        - Payer Name
        - Currency
        - Supplier Name
        - Total Amount of the Invoice
        - Total VAT or Tax
        
        Invoice text:
        {sample_text}
        
        Return only valid JSON format with the above fields as keys.
        """
        
        response = openai.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are an expert at extracting structured data from invoices. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=1000
        )
        
        result = response.choices[0].message.content
        print("✅ Data extraction successful")
        print("Sample extraction result:")
        print(result)
        
        # Try to parse JSON
        try:
            data = json.loads(result)
            print("✅ JSON parsing successful")
            return True
        except json.JSONDecodeError:
            print("❌ JSON parsing failed")
            return False
            
    except Exception as e:
        print(f"❌ Data extraction failed: {e}")
        return False

def main():
    """Run all tests"""
    print("🔍 Testing Invoice Data Extractor Setup (No OCR)\n")
    
    tests = [
        ("Package Imports", test_imports),
        ("OpenAI API Key", test_openai_key),
        ("PDF Processing Libraries", test_pdf_processing),
        ("Excel File Creation", test_excel_creation),
        ("Data Extraction", test_data_extraction)
    ]
    
    results = []
    for test_name, test_func in tests:
        print(f"\n--- {test_name} ---")
        results.append(test_func())
    
    print(f"\n{'='*50}")
    if all(results):
        print("🎉 All tests passed! Your setup is ready.")
        print("\nTo run the app:")
        print("  streamlit run app_no_ocr.py")
        print("\nNote: OCR will work on Streamlit Cloud!")
    else:
        print("❌ Some tests failed. Please fix the issues above.")
        print("\nCommon solutions:")
        print("1. Install missing packages: pip install -r requirements_no_ocr.txt")
        print("2. Set your OpenAI API key in .env file")

if __name__ == "__main__":
    main()
