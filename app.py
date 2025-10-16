import streamlit as st
import pandas as pd
import openai
import os
from dotenv import load_dotenv
import PyPDF2
import pdfplumber
import pytesseract
from PIL import Image
import fitz  # PyMuPDF
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()

# Configure page
st.set_page_config(
    page_title="Invoice Data Extractor",
    page_icon="📄",
    layout="wide"
)

# Initialize OpenAI
def initialize_openai():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        st.error("Please set your OPENAI_API_KEY in the .env file")
        st.stop()
    openai.api_key = api_key
    return openai

# Extract text from PDF (handles both searchable and scanned PDFs)
def extract_text_from_pdf(pdf_file):
    text = ""
    
    try:
        # First try with pdfplumber (better for searchable PDFs)
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        
        # If no text extracted, try OCR with PyMuPDF
        if not text.strip():
            pdf_document = fitz.open(stream=pdf_file.read(), filetype="pdf")
            for page_num in range(pdf_document.page_count):
                page = pdf_document[page_num]
                # Convert page to image
                pix = page.get_pixmap()
                img_data = pix.tobytes("png")
                image = Image.open(io.BytesIO(img_data))
                
                # Use OCR to extract text
                page_text = pytesseract.image_to_string(image, config='--psm 6')
                text += page_text + "\n"
            
            pdf_document.close()
    
    except Exception as e:
        st.error(f"Error extracting text from PDF: {str(e)}")
        return ""
    
    return text

# Extract data using OpenAI GPT
def extract_invoice_data(text, openai_client):
    prompt = f"""
    You are an expert at extracting structured data from invoices. Extract the following information from the provided invoice text and return it in JSON format.
    If any information is not found, use "N/A" as the value.
    
    IMPORTANT: Look for these fields with various possible names/abbreviations:
    - PO Number (could be: PO No, Purchase Order, P.O. Number, etc.)
    - Item Code (could be: Item No, Product Code, SKU, Part Number, etc.)
    - Description (could be: Product Description, Item Description, etc.)
    - UOM (could be: Unit of Measure, Unit, U/M, etc.)
    - Quantity (could be: Qty, Amount, etc.)
    - Lot Number (could be: Lot No, Batch Number, Batch No, etc.)
    - Expiry Date (could be: Exp Date, Expiration Date, etc.)
    - Mfg Date (could be: Manufacturing Date, Mfg Date, Production Date, etc.)
    - Invoice No (could be: Invoice Number, Inv No, etc.)
    - Unit Price (could be: Price per Unit, Unit Cost, etc.)
    - Total Price (could be: Line Total, Item Total, etc.)
    - Country (could be: Origin Country, Country of Origin, etc.)
    - HS Code (could be: HSN Code, Tariff Code, etc.)
    - Date of Invoice (could be: Invoice Date, Date, etc.)
    - Customer No (could be: Customer Number, Customer ID, Client No, etc.)
    - Payer Name (could be: Payer, Bill To, etc.)
    - Currency (could be: Curr, etc.)
    - Supplier Name (could be: Vendor Name, Supplier, Company Name, etc.)
    - Total Amount of the Invoice (could be: Grand Total, Total Amount, Invoice Total, etc.)
    - Total VAT or Tax (could be: VAT, Tax, Tax Amount, VAT Amount, etc.)
    
    Invoice text:
    {text}
    
    Return only valid JSON format with the above fields as keys. Use the exact field names provided above.
    """
    
    try:
        response = openai_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are an expert at extracting structured data from invoices. Always return valid JSON format with the exact field names provided."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=2000
        )
        
        return response.choices[0].message.content
    except Exception as e:
        st.error(f"Error calling OpenAI API: {str(e)}")
        return None

# Create Excel file with formatting
def create_excel_file(data_list, filename="extracted_invoice_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    
    # Define headers
    headers = [
        "PO Number", "Item Code", "Description", "UOM", "Quantity", "Lot Number",
        "Expiry Date", "Mfg Date", "Invoice No", "Unit Price", "Total Price",
        "Country", "HS Code", "Date of Invoice", "Customer No", "Payer Name",
        "Currency", "Supplier Name", "Total Amount of the Invoice", "Total VAT or Tax"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, data in enumerate(data_list, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=data.get(header, "N/A"))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(filename)
    return filename

# Main Streamlit app
def main():
    st.title("📄 Invoice Data Extractor")
    st.markdown("Upload one or multiple PDF files to extract invoice data using AI")
    
    # Initialize OpenAI
    openai_client = initialize_openai()
    
    # File upload
    uploaded_files = st.file_uploader(
        "Choose PDF files",
        type=['pdf'],
        accept_multiple_files=True,
        help="Upload one or multiple PDF files (searchable or scanned)"
    )
    
    if uploaded_files:
        st.success(f"Uploaded {len(uploaded_files)} file(s)")
        
        # Process files
        if st.button("Extract Data", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            all_extracted_data = []
            
            for i, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"Processing {uploaded_file.name}...")
                
                # Extract text from PDF
                text = extract_text_from_pdf(uploaded_file)
                
                if text.strip():
                    # Extract data using OpenAI
                    extracted_data = extract_invoice_data(text, openai_client)
                    
                    if extracted_data:
                        try:
                            # Parse JSON response
                            import json
                            data_dict = json.loads(extracted_data)
                            data_dict['Source File'] = uploaded_file.name
                            all_extracted_data.append(data_dict)
                        except json.JSONDecodeError:
                            st.warning(f"Could not parse data from {uploaded_file.name}")
                    else:
                        st.warning(f"Could not extract data from {uploaded_file.name}")
                else:
                    st.warning(f"No text could be extracted from {uploaded_file.name}")
                
                progress_bar.progress((i + 1) / len(uploaded_files))
            
            if all_extracted_data:
                # Create DataFrame
                df = pd.DataFrame(all_extracted_data)
                
                # Display extracted data
                st.success("Data extracted successfully!")
                st.dataframe(df, use_container_width=True)
                
                # Create and download Excel file
                excel_filename = f"extracted_invoice_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                create_excel_file(all_extracted_data, excel_filename)
                
                with open(excel_filename, "rb") as file:
                    st.download_button(
                        label="Download Excel File",
                        data=file.read(),
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                # Clean up temporary file
                os.remove(excel_filename)
            else:
                st.error("No data could be extracted from any of the files")

if __name__ == "__main__":
    main()
